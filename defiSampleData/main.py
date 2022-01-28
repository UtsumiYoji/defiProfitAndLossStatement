from datetime import datetime
import random

import openpyxl

def main():
    # 初期価格と初期所持量を定義
    cryptoPrice = [4000000, 270000, 67.86, 0.7238625, 909.72]
    cryptoAmount = [0.25, 3.3, 117, 13600, 920]

    # ブックを新規作成
    wb = openpyxl.Workbook()
    sh = wb.active

    # シートの名前をつけてあげる
    sh.title = 'Sheet'

    # 初期保持量を出力
    for j in range(len(cryptoPrice)):
        sh.cell(j+1, len(cryptoPrice)+6).value = 'crypto' + str(j+1)
        sh.cell(j+1, len(cryptoPrice)+7).value = cryptoPrice[j]
        sh.cell(j+1, len(cryptoPrice)+8).value = cryptoAmount[j]
        sh.cell(j+1, len(cryptoPrice)+9).value = cryptoAmount[j] * cryptoPrice[j]

    # 100回ループする
    for i in range(100):
        # 通貨の数ループする
        for j in range(len(cryptoPrice)):
            # 変動率を取得して変動後の価格に置き換え
            changePar = random.uniform(-0.1, 0.1)
            cryptoPrice[j] += (cryptoPrice[j] * changePar)

            # 価格を記録
            sh.cell(i+1, j+1).value = cryptoPrice[j]

        # 取引する通貨を2つ選ぶ
        cryptoA = random.randint(0, len(cryptoPrice)-1)
        while True:
            cryptoB = random.randint(0, len(cryptoPrice)-1)
            if not cryptoA == cryptoB:
                break
        
        # 取引量を考える
        changeAmount = random.uniform(0.1, 0.5)

        # 取引量と取引通貨を記録
        sh.cell(i+1, len(cryptoPrice)+2).value = changeAmount
        sh.cell(i+1, len(cryptoPrice)+3).value = cryptoA+1
        sh.cell(i+1, len(cryptoPrice)+4).value = cryptoB+1

        # 取引を行う
        cryptoAmount[cryptoA] -= changeAmount * cryptoAmount[cryptoA]
        cryptoAmount[cryptoB] += (changeAmount * cryptoAmount[cryptoA] * cryptoPrice[cryptoA])/cryptoPrice[cryptoB]

        # 最終保持量を出力
        for j in range(len(cryptoPrice)):
            sh.cell(j+1, len(cryptoPrice)+10).value = cryptoPrice[j]
            sh.cell(j+1, len(cryptoPrice)+11).value = cryptoAmount[j]
            sh.cell(j+1, len(cryptoPrice)+12).value = cryptoAmount[j] * cryptoPrice[j]
            
    # ブックを保存
    wb.save('./' + datetime.now().strftime('%Y%m%d%H%M%S') + ' サンプルデータ.xlsx')

if __name__ == '__main__':
    main()