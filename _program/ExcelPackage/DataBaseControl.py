import openpyxl

class DataBaseControl:
    def __init__(self) -> None:
        self.wb = openpyxl.open('./database.xlsx')
        self.sh = self.wb['database']
    
    def __del__(self) -> None:
        self.wb.close()

    def getSymbol(self, chain) -> str:
        for i in range(1, self.sh.max_row+1):
            if chain == self.sh.cell(i, 1).value:
                return self.sh.cell(i, 2).value
    
    def GetContract(self, chain) -> str:
        for i in range(1, self.sh.max_row+1):
            if chain == self.sh.cell(i, 1).value:
                return self.sh.cell(i, 3).value