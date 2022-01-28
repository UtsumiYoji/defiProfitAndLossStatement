import swapCalculation
import openpyxl

def main():
    # インタンスを作成する
    swapCalIns = swapCalculation.swapCalculation()

    # サンプルデータを読み込む
    wb = openpyxl.load_workbook('./20220128212354 サンプルデータ.xlsx')
    sh = wb['Sheet']

    # 初期保有量を与える
    tokenName = ['BTC', 'ETH', 'RIP', 'DOGE', 'CAKE']
    for row in range(len(tokenName)):
        swapCalIns.addAmount(tokenName[row], sh.cell(row+1, 15).value, sh.cell(row+1, 14).value)
    
    # 下までループ
    for row in range(sh.max_row):
        # 交換に使ったトークンを記録
        a = sh.cell(row+1, 8).value -1
        b = sh.cell(row+1, 10).value -1

        swapCalIns.addSwapData(
            tokenName[a], sh.cell(row+1, 9).value, tokenName[b], sh.cell(row+1, 11).value, sh.cell(row+1, b+1).value
        )
    
    # 現在の状況をプリント
    for i in range(len(tokenName)):
        print(tokenName[i]+':', end=' ')
        print('保持数:'+str(swapCalIns.amount[tokenName[i]]), end=' ')
        print('平均取得単価:'+str(swapCalIns.averageCost[tokenName[i]]), end=' ')
        print('現在の価格:'+str(sh.cell(i+1, 17).value))

    print('確定総損益：' + str(swapCalIns.profit))

if __name__ == '__main__':
    main()